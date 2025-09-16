import csv, io
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.contrib.auth import login as auth_login
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.db.models import Q, Sum, F, DecimalField, ExpressionWrapper, Avg, Min, Max
import math
from decimal import Decimal
from .constants import STATUSES, CATEGORIES, CO_MANAGER_GROUP
from .models import Product, Variant, ProductImage
from .forms import ProductForm, VariantForm, ImportFileForm
from .csv_sync import schedule_csv_sync
from django.utils.text import slugify
import zipfile, json, os

@login_required
def dashboard(request):
    # Persist filters in session
    session_key = 'dashboard_filters'
    clear = request.GET.get('clear')
    if clear:
        request.session.pop(session_key, None)
    keys = ('q', 'status', 'cat', 'sort', 'archived')
    has_any = any(k in request.GET for k in keys)
    if has_any:
        filters = {
            'q': (request.GET.get('q') or '').strip(),
            'status': (request.GET.get('status') or '').strip(),
            'cat': (request.GET.get('cat') or '').strip(),
            'sort': (request.GET.get('sort') or '').strip(),
            'archived': (request.GET.get('archived') or '').strip(),
        }
        request.session[session_key] = filters
    else:
        filters = request.session.get(session_key, {'q':'','status':'','cat':'','sort':'','archived':''})

    q = filters.get('q','')
    status = filters.get('status','')
    cat = filters.get('cat','')
    sort = filters.get('sort','')
    archived_flag = (filters.get('archived','') in ('1','true','yes'))
    # Co-managers cannot view archived
    if (
        request.user.is_authenticated
        and request.user.is_staff
        and request.user.groups.filter(name=CO_MANAGER_GROUP).exists()
    ):
        archived_flag = False
    products_qs = Product.objects.prefetch_related('variants')
    if q:
        products_qs = products_qs.filter(
            Q(name__icontains=q) |
            Q(brand__icontains=q) |
            Q(category__icontains=q) |
            Q(main_sku__iexact=q) |
            Q(main_sku__icontains=q) |
            Q(variants__variant_sku__iexact=q) |
            Q(variants__variant_sku__icontains=q)
        )
    if status and status in STATUSES:
        products_qs = products_qs.filter(variants__status=status)
    if cat:
        products_qs = products_qs.filter(category=cat)
    # Archived filter: archived=1 shows ONLY archived; default shows ONLY active
    if archived_flag:
        products_qs = products_qs.filter(archived=True)
    else:
        products_qs = products_qs.filter(archived=False)
    # Sorting
    sort_map = {
        'created_desc': '-id',
        'created_asc': 'id',
        'name_az': 'name',
        'name_za': '-name',
        'brand_az': 'brand',
        'brand_za': '-brand',
        'sku_az': 'main_sku',
        'sku_za': '-main_sku',
        'category_az': 'category',
        'category_za': '-category',
    }
    order_by = sort_map.get(sort, '-id')
    products = products_qs.order_by(order_by).distinct()[:100]
    # Build category suggestions from constants + DB
    db_cats = list(Product.objects.values_list('category', flat=True).distinct())
    cat_suggestions = sorted({*(c for c in CATEGORIES), *(c for c in db_cats if c)})
    return render(request, 'inventory/dashboard.html', {
        'products': products,
        'q': q,
        'status': status,
        'cat': cat,
        'sort': sort,
        'statuses': STATUSES,
        'categories': cat_suggestions,
        'show_archived': archived_flag,
    })

@login_required
def home(request):
    # High-level KPIs; profit only from Sold variants
    sold_qs = Variant.objects.filter(status='Sold')
    listed_qs = Variant.objects.filter(status='Listed')
    unsold_qs = Variant.objects.exclude(status='Sold')
    total_profit = sold_qs.aggregate(total=Sum('profit'))['total'] or Decimal('0')
    total_net = sold_qs.aggregate(total=Sum('net'))['total'] or Decimal('0')
    # Stock-level sums (qty-aware)
    sum_price_qty = ExpressionWrapper(F('price') * F('qty'), output_field=DecimalField(max_digits=12, decimal_places=2))
    sum_cost_qty = ExpressionWrapper(F('cost') * F('qty'), output_field=DecimalField(max_digits=12, decimal_places=2))
    stock_list_value = unsold_qs.aggregate(total=Sum(sum_price_qty))['total'] or Decimal('0')
    stock_cost_value = unsold_qs.aggregate(total=Sum(sum_cost_qty))['total'] or Decimal('0')
    listed_list_value = listed_qs.aggregate(total=Sum(sum_price_qty))['total'] or Decimal('0')

    totals = {
        'products': Product.objects.count(),
        'variants': Variant.objects.count(),
        'sold': sold_qs.aggregate(q=Sum('qty'))['q'] or 0,
        'listed': listed_qs.aggregate(q=Sum('qty'))['q'] or 0,
        'draft': Variant.objects.filter(status='Draft').aggregate(q=Sum('qty'))['q'] or 0,
    }
    # Additional metrics for richer dashboard
    total_qty = Variant.objects.aggregate(q=Sum('qty'))['q'] or 0
    sell_through_pct = int(((totals['sold'] or 0) / total_qty) * 100) if total_qty else 0
    avg_list_price = unsold_qs.aggregate(a=Avg('price'))['a'] or Decimal('0')
    avg_cost = Variant.objects.aggregate(a=Avg('cost'))['a'] or Decimal('0')
    categories_count = Product.objects.values('category').distinct().count()
    brands_count = Product.objects.exclude(brand='').values('brand').distinct().count()
    avg_profit_per_sale = (total_profit / (totals['sold'] or 1)) if totals['sold'] else Decimal('0')
    unsold_qty = max(0, (total_qty or 0) - (totals['sold'] or 0))
    # Status distribution for donut
    status_counts = {
        'Sold': int(totals['sold'] or 0),
        'Listed': int(totals['listed'] or 0),
        'Draft': int(totals['draft'] or 0),
    }
    other_count = max(0, int(total_qty) - sum(status_counts.values()))
    if other_count:
        status_counts['Other'] = other_count

    # Prepare conic-gradient segments for Inventory Mix donut
    mix_palette = {
        'Sold': '#10b981',
        'Listed': '#27A7FF',
        'Draft': '#f59e0b',
        'Other': '#a78bfa',
    }
    status_mix_segments = []
    acc = 0
    if total_qty and total_qty > 0:
        for label in ['Sold', 'Listed', 'Draft', 'Other']:
            v = int(status_counts.get(label, 0) or 0)
            if v <= 0:
                continue
            pct = int(round((v / int(total_qty)) * 100))
            seg = {
                'label': label,
                'value': v,
                'color': mix_palette[label],
                'frm': acc,
                'to': min(100, acc + pct),
            }
            status_mix_segments.append(seg)
            acc = seg['to']
        # Ensure last segment ends at 100% to avoid visual gap from rounding
        if status_mix_segments:
            status_mix_segments[-1]['to'] = 100
    else:
        # No data -> show a muted ring
        status_mix_segments = []
    # Milestones (simple defaults)
    milestone_targets = [Decimal('100'), Decimal('500'), Decimal('1000'), Decimal('5000')]
    next_target = None
    for m in milestone_targets:
        if total_profit < m:
            next_target = m
            break
    progress_pct = int((total_profit / next_target * 100)) if next_target and next_target > 0 else 100
    remaining_to_target = (next_target - total_profit) if next_target else Decimal('0')
    if remaining_to_target < 0:
        remaining_to_target = Decimal('0')
    # Top categories by sold count
    # Top categories with percentage bars
    raw_top_categories = list(
        sold_qs.select_related('product')
        .values('product__category')
        .annotate(count=Sum('qty'))
        .order_by('-count')[:5]
    )
    max_cat = max([row['count'] for row in raw_top_categories], default=0)
    top_categories = [
        {
            'category': row['product__category'],
            'count': row['count'],
            'pct': int((row['count'] / max_cat * 100)) if max_cat else 0,
        }
        for row in raw_top_categories
    ]
    top_category_name = raw_top_categories[0]['product__category'] if raw_top_categories else ''

    # Top brands by units sold and profit
    top_brands = list(
        sold_qs.select_related('product')
        .values('product__brand')
        .annotate(count=Sum('qty'), profit=Sum('profit'))
        .order_by('-count')[:5]
    )
    best_profit_brand = (
        sold_qs.select_related('product')
        .values('product__brand')
        .annotate(p=Sum('profit'))
        .order_by('-p')
        .first()
    )

    # Top locations by units (all variants)
    top_locations = list(
        Variant.objects.values('location').annotate(q=Sum('qty')).order_by('-q')[:5]
    )
    max_loc = max([row['q'] for row in top_locations], default=0)

    avg_margin = sold_qs.aggregate(m=Avg('margin'))['m'] or Decimal('0')
    recent = Variant.objects.select_related('product').order_by('-id')[:6]
    ctx = {
        'total_profit': total_profit,
        'total_net': total_net,
        'totals': totals,
        'total_qty': total_qty,
        'sell_through_pct': sell_through_pct,
        'avg_list_price': avg_list_price,
        'avg_cost': avg_cost,
        'categories_count': categories_count,
        'brands_count': brands_count,
        'avg_profit_per_sale': avg_profit_per_sale,
        'unsold_qty': unsold_qty,
        'status_counts': status_counts,
        'status_mix_segments': status_mix_segments,
        'next_target': next_target,
        'progress_pct': progress_pct,
        'remaining_to_target': remaining_to_target,
        'top_categories': top_categories,
        'top_brands': top_brands,
        'top_locations': top_locations,
        'top_locations_max': max_loc,
        'top_category_name': top_category_name,
        'best_profit_brand': best_profit_brand,
        'avg_margin': avg_margin,
        'stock_list_value': stock_list_value,
        'stock_cost_value': stock_cost_value,
        'listed_list_value': listed_list_value,
        'recent': recent,
    }
    return render(request, 'inventory/home.html', ctx)

@login_required
def bulk_update(request):
    if request.method != 'POST':
        return redirect('inventory:dashboard')
    ids = request.POST.getlist('ids')
    set_status = request.POST.get('set_status', '').strip()
    set_location = request.POST.get('set_location', '').strip()
    set_category = request.POST.get('set_category', '').strip()

    if not ids:
        messages.error(request, 'No items selected.')
        return redirect('inventory:dashboard')

    # Update selected products and their variants
    products = Product.objects.filter(pk__in=ids)
    updated = 0
    if set_category:
        updated += products.update(category=set_category)
    if set_status or set_location:
        vqs = Variant.objects.filter(product__in=products)
        update_kwargs = {}
        if set_status and set_status in STATUSES:
            update_kwargs['status'] = set_status
        if set_location:
            update_kwargs['location'] = set_location
        if update_kwargs:
            updated += vqs.update(**update_kwargs)

    messages.success(request, f'Updated {updated} fields on selected items.')
    schedule_csv_sync()
    return redirect('inventory:dashboard')

@login_required
def product_create(request):
    if request.method == 'POST':
        pform = ProductForm(request.POST)
        vform = VariantForm(request.POST, request.FILES)
        if pform.is_valid() and vform.is_valid():
            product = pform.save()
            variant = vform.save(commit=False)
            variant.product = product
            variant.save()
            # handle multiple images
            files = request.FILES.getlist('new_images') or request.FILES.getlist('images')
            for f in files:
                ProductImage.objects.create(variant=variant, image=f)
            messages.success(request, 'Product created successfully.')
            schedule_csv_sync()
            return redirect('inventory:product_detail', pk=product.pk)
        else:
            pass
    else:
        pform = ProductForm()
        vform = VariantForm()
    db_cats = list(Product.objects.values_list('category', flat=True).distinct())
    cat_suggestions = sorted({*(c for c in CATEGORIES), *(c for c in db_cats if c)})
    return render(request, 'inventory/product_form.html', {'pform': pform, 'vform': vform, 'categories': cat_suggestions})

@login_required
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'inventory/product_detail.html', {'product': product})

@login_required
def ebay_search(request):
    """Minimal JSON proxy for eBay Browse search.
    GET /inventory/api/ebay/search?q=shoes&limit=5
    """
    q = (request.GET.get('q') or '').strip()
    limit = int(request.GET.get('limit') or 10)
    if not q:
        return JsonResponse({'error': 'Missing q parameter'}, status=400)
    # Lazy import so missing optional dependency (requests) doesn't block app startup
    from .ebay import get_client
    client = get_client()
    if not client:
        return JsonResponse({'error': 'eBay not configured; set EBAY_* settings and enable EBAY_ENABLED=1'}, status=503)
    try:
        data = client.search(q=q, limit=limit)
        items = data.get('itemSummaries') or []
        stats = client.summarize_prices(items)
        return JsonResponse({'q': q, 'count': len(items), 'stats': stats, 'items': items})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=502)

@login_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product updated.')
            schedule_csv_sync()
            return redirect('inventory:product_detail', pk=product.pk)
        else:
            pass
    else:
        form = ProductForm(instance=product)
    db_cats = list(Product.objects.values_list('category', flat=True).distinct())
    cat_suggestions = sorted({*(c for c in CATEGORIES), *(c for c in db_cats if c)})
    return render(request, 'inventory/product_edit.html', {'form': form, 'product': product, 'categories': cat_suggestions})

@login_required
def variant_edit(request, pk):
    variant = get_object_or_404(Variant, pk=pk)
    if request.method == 'POST':
        form = VariantForm(request.POST, request.FILES, instance=variant)
        if form.is_valid():
            variant = form.save()
            files = request.FILES.getlist('new_images') or request.FILES.getlist('images')
            for f in files:
                ProductImage.objects.create(variant=variant, image=f)
            messages.success(request, 'Variant updated.')
            schedule_csv_sync()
            # Stay on edit page instead of closing to product detail
            return redirect('inventory:variant_edit', pk=variant.pk)
        else:
            pass
    else:
        form = VariantForm(instance=variant)
    return render(request, 'inventory/variant_form.html', {'form': form, 'variant': variant})

@login_required
def variant_create(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = VariantForm(request.POST, request.FILES)
        if form.is_valid():
            variant = form.save(commit=False)
            variant.product = product
            variant.save()
            files = request.FILES.getlist('new_images') or request.FILES.getlist('images')
            for f in files:
                ProductImage.objects.create(variant=variant, image=f)
            messages.success(request, 'Variant added.')
            schedule_csv_sync()
            return redirect('inventory:product_detail', pk=product.pk)
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = VariantForm()
    return render(request, 'inventory/variant_create.html', {'form': form, 'product': product})

@login_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        name = product.name
        product.delete()
        messages.success(request, f'Product "{name}" deleted.')
        schedule_csv_sync()
        return redirect('inventory:dashboard')
    return render(request, 'inventory/product_confirm_delete.html', {'product': product})

@login_required
def variant_delete(request, pk):
    variant = get_object_or_404(Variant, pk=pk)
    product = variant.product
    if request.method == 'POST':
        sku = variant.variant_sku
        variant.delete()
        messages.success(request, f'Variant {sku or ""} deleted.')
        schedule_csv_sync()
        return redirect('inventory:product_detail', pk=product.pk)
    return render(request, 'inventory/variant_confirm_delete.html', {'variant': variant})

@login_required
def image_delete(request, pk):
    img = get_object_or_404(ProductImage, pk=pk)
    variant = img.variant
    if request.method == 'POST':
        img.delete()
        messages.success(request, 'Image deleted.')
        schedule_csv_sync()
        return redirect('inventory:variant_edit', pk=variant.pk)
    return redirect('inventory:variant_edit', pk=variant.pk)

@login_required
def product_archive(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.archived = True
        product.save(update_fields=['archived'])
        messages.success(request, f'Archived {product.name}.')
        schedule_csv_sync()
        return redirect('inventory:dashboard')
    return redirect('inventory:product_detail', pk=pk)

@login_required
def product_unarchive(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.archived = False
        product.save(update_fields=['archived'])
        messages.success(request, f'Restored {product.name}.')
        schedule_csv_sync()
        return redirect('inventory:dashboard')
    return redirect('inventory:product_detail', pk=pk)

def store_index(request):
    # Secret development storefront: shows publicly visible listed items
    q = (request.GET.get('q') or '').strip()
    sort = (request.GET.get('sort') or '').strip()
    # Integer price bounds in increments of 5 for UI; convert to Decimal for DB filter
    def _parse_int(v):
        try:
            if v in (None, ''):
                return None
            return int(v)
        except Exception:
            return None
    min_price_int = _parse_int(request.GET.get('min_price'))
    max_price_int = _parse_int(request.GET.get('max_price'))
    items = Variant.objects.select_related('product').prefetch_related('images').filter(
        status='Listed', product__archived=False
    )
    if q:
        items = items.filter(
            Q(product__name__icontains=q) |
            Q(product__brand__icontains=q) |
            Q(product__category__icontains=q) |
            Q(variant_sku__icontains=q) |
            Q(size__icontains=q) |
            Q(colour__icontains=q)
        )
    if min_price_int is not None:
        items = items.filter(price__gte=Decimal(min_price_int))
    if max_price_int is not None:
        items = items.filter(price__lte=Decimal(max_price_int))
    # Sorting
    order_map = {
        'price_asc': 'price',
        'price_desc': '-price',
        'newest': '-id',
    }
    items = items.order_by(order_map.get(sort, '-id'))
    items = items[:120]
    raw_bounds = Variant.objects.filter(status='Listed', product__archived=False).aggregate(mn=Min('price'), mx=Max('price'))
    mn = raw_bounds['mn'] or Decimal('0')
    mx = raw_bounds['mx'] or Decimal('0')
    # Round to nearest 5s for slider defaults
    mn5 = int(math.floor(float(mn) / 5.0) * 5)
    mx5 = int(math.ceil(float(mx) / 5.0) * 5)
    return render(request, 'store.html', {
        'items': items,
        'q': q,
        'sort': sort,
        'min_price': min_price_int,
        'max_price': max_price_int,
        'bounds': {'mn': mn5, 'mx': mx5},
    })

def _cart_get(request):
    return request.session.get('store_cart', {})

def _cart_set(request, cart):
    request.session['store_cart'] = cart
    request.session.modified = True

def store_product(request, vid):
    v = get_object_or_404(Variant.objects.select_related('product').prefetch_related('images'), pk=vid, status='Listed', product__archived=False)
    if request.method == 'POST':
        # Add to cart
        qty = max(1, int(request.POST.get('qty', '1') or '1'))
        cart = _cart_get(request)
        cart[str(v.id)] = cart.get(str(v.id), 0) + qty
        _cart_set(request, cart)
        messages.success(request, f'Added {qty} × {v.product.name} ({v.size}) to cart.')
        return redirect('store_cart')
    return render(request, 'store_detail.html', {'v': v})

def store_cart(request):
    cart = _cart_get(request)
    ids = [int(k) for k in cart.keys()]
    variants = Variant.objects.select_related('product').prefetch_related('images').filter(id__in=ids, status='Listed', product__archived=False)
    items = []
    subtotal = Decimal('0')
    for v in variants:
        qty = int(cart.get(str(v.id), 0))
        line = {'v': v, 'qty': qty, 'line_total': v.price * qty}
        items.append(line)
        subtotal += line['line_total']
    if request.method == 'POST':
        action = request.POST.get('action')
        vid = request.POST.get('vid')
        if action == 'update' and vid:
            qty = max(0, int(request.POST.get('qty', '1') or '1'))
            if qty == 0:
                cart.pop(vid, None)
            else:
                cart[vid] = qty
            _cart_set(request, cart)
            return redirect('store_cart')
        elif action == 'clear':
            _cart_set(request, {})
            return redirect('store_cart')
        elif action == 'checkout':
            return redirect('store_checkout')
    return render(request, 'store_cart.html', {'items': items, 'subtotal': subtotal})

def store_checkout(request):
    cart = _cart_get(request)
    ids = [int(k) for k in cart.keys()]
    variants = Variant.objects.select_related('product').filter(id__in=ids, status='Listed', product__archived=False)
    subtotal = Decimal('0')
    for v in variants:
        subtotal += (v.price * (cart.get(str(v.id)) or 0))
    if request.method == 'POST':
        # Simulate order placed
        _cart_set(request, {})
        messages.success(request, 'Order placed! (dev preview — no payment processed)')
        return redirect('store_index')
    return render(request, 'store_checkout.html', {'subtotal': subtotal, 'count': sum(cart.values())})

@login_required
def settings_view(request):
    pwd_form = PasswordChangeForm(user=request.user)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'change_password':
            pwd_form = PasswordChangeForm(user=request.user, data=request.POST)
            if pwd_form.is_valid():
                user = pwd_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, 'Password updated successfully.')
                return redirect('inventory:settings')
            else:
                messages.error(request, 'Please correct the errors in the password form.')
        elif action == 'create_comanager':
            # Allow any staff user (admin) to create co-managers
            if not request.user.is_staff:
                messages.error(request, 'You do not have permission to create co-managers.')
                return redirect('inventory:settings')
            username = (request.POST.get('username') or '').strip()
            p1 = request.POST.get('password1') or ''
            p2 = request.POST.get('password2') or ''
            if not username or not p1:
                messages.error(request, 'Username and password are required for co-manager.')
            elif p1 != p2:
                messages.error(request, 'Passwords do not match for co-manager.')
            elif User.objects.filter(username=username).exists():
                messages.error(request, 'Username is already taken.')
            else:
                u = User.objects.create_user(username=username, password=p1)
                u.is_staff = True
                u.save()
                group, _ = Group.objects.get_or_create(name=CO_MANAGER_GROUP)
                u.groups.add(group)
                messages.success(request, f'Co-manager account "{username}" created.')
                return redirect('inventory:settings')
    return render(request, 'inventory/settings.html', {
        'pwd_form': pwd_form,
    })

def signup(request):
    # Re-enabled: allow signup regardless of existing users
    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip()
        pwd1 = request.POST.get('password1') or ''
        pwd2 = request.POST.get('password2') or ''
        if not username or not pwd1:
            messages.error(request, 'Username and password are required.')
        elif pwd1 != pwd2:
            messages.error(request, 'Passwords do not match.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'Username already taken.')
        else:
            user = User.objects.create_user(username=username, password=pwd1)
            auth_login(request, user)
            return redirect('inventory:home')
    return render(request, 'auth/signup.html')

def import_products(request):
    if request.method == 'POST':
        form = ImportFileForm(request.POST, request.FILES)
        if form.is_valid():
            f = form.cleaned_data['file']
            name = f.name.lower()
            try:
                if name.endswith('.csv'):
                    decoded = f.read().decode('utf-8')
                    # Robust delimiter detection based on header column count
                    lines = [ln for ln in decoded.splitlines() if ln.strip()]
                    header = lines[0] if lines else ''
                    candidates = ['\t', ',', ';', '|']
                    best = max(candidates, key=lambda d: len(header.split(d)))
                    if len(header.split(best)) <= 1:
                        best = ','
                    reader = csv.DictReader(io.StringIO(decoded), delimiter=best)
                    count = _ingest_rows(reader)
                elif name.endswith('.xlsx'):
                    try:
                        import openpyxl
                    except ImportError:
                        messages.error(request, 'XLSX import requires openpyxl. Install it and try again.')
                        return redirect('inventory:import_products')
                    wb = openpyxl.load_workbook(f)
                    ws = wb.active
                    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    rows = (dict(zip(headers, [cell.value for cell in row])) for row in ws.iter_rows(min_row=2))
                    count = _ingest_rows(rows)
                else:
                    messages.error(request, 'Unsupported file type. Please upload CSV or XLSX.')
                    return redirect('inventory:import_products')
                messages.success(request, f'Imported {count} rows successfully.')
                return redirect('inventory:dashboard')
            except Exception as e:
                messages.error(request, f'Import failed: {e}')
                return redirect('inventory:import_products')
    else:
        form = ImportFileForm()
    return render(request, 'inventory/import.html', {'form': form})

def _ingest_rows(rows):
    count = 0
    for row in rows:
        # Normalize header keys: trim whitespace and unify casing
        norm = {}
        for k, v in row.items():
            kk = (k or '')
            if isinstance(kk, str):
                kk = kk.strip()
            norm[kk] = v
        row = norm
        # Accept flexible headers and synonyms
        def get_val(keys, default=''):
            for k in keys:
                if k in row and row.get(k) is not None:
                    v = row.get(k)
                    return v.strip() if isinstance(v, str) else v
            return default

        # Core fields
        name = (get_val(['Product Name', 'Title', 'Name'], '') or '').strip()
        if not name:
            continue
        brand = (get_val(['Brand'], '') or '')
        category = (get_val(['Category'], 'Clothing') or 'Clothing')

        # Optional main SKU handling (support Master/Main); zero-pad numeric
        main_sku_in = str(get_val(['Main SKU', 'Master SKU'], '') or '').strip()
        if main_sku_in.isdigit():
            main_sku_in = f"{int(main_sku_in):03d}"

        product = None
        if main_sku_in:
            product = Product.objects.filter(main_sku=main_sku_in).first()
        if product is None:
            product = Product.objects.create(name=name, brand=brand, category=category, main_sku=main_sku_in or '')

        variant_sku_in = (get_val(['Variant SKU', 'SKU Variant'], '') or '').strip()
        variant = None
        if variant_sku_in:
            variant = Variant.objects.filter(variant_sku=variant_sku_in).first()
        if variant is None:
            variant = Variant(product=product)
            if variant_sku_in:
                variant.variant_sku = variant_sku_in
        # assign fields
        variant.size = (get_val(['Size'], '') or '')[:40]
        variant.condition = (get_val(['Condition'], 'Good') or 'Good')
        variant.colour = (get_val(['Colour', 'Color'], '') or '')
        try:
            variant.qty = int(get_val(['Qty', 'Quantity'], 1) or 1)
        except Exception:
            variant.qty = 1
        variant.location = (get_val(['Location', 'Location/Bin', 'Bin', 'Shelf'], 'Spare Room') or 'Spare Room')
        variant.status = (get_val(['Status'], 'Draft') or 'Draft')
        # numeric and date parsing
        def parse_money(val):
            if val is None: return 0
            s = str(val).replace('£','').strip()
            return float(s or 0)
        def parse_date(val):
            if not val: return datetime.now().date()
            if isinstance(val, datetime): return val.date()
            for fmt in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y'):
                try: return datetime.strptime(str(val), fmt).date()
                except: pass
            return datetime.now().date()

        variant.date = parse_date(get_val(['Date', 'Purchase Date']))
        variant.cost = parse_money(get_val(['Cost', 'Purchase Price', 'Buy Price']))
        variant.price = parse_money(get_val(['Price', 'Listed Price', 'Sale Price']))
        variant.fees = parse_money(get_val(['Fees', 'Estimated Fees', 'Platform Fees']))
        variant.save()
        count += 1
    return count

@login_required
def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products.csv"'
    writer = csv.writer(response)
    writer.writerow(['Main SKU','Variant SKU','Product Name','Brand','Category','Size','Condition','Colour','Date','Cost','Price','Fees','Net','Profit','Margin','Qty','Location','Status'])
    for v in Variant.objects.select_related('product').all():
        writer.writerow([
            v.product.main_sku, v.variant_sku, v.product.name, v.product.brand, v.product.category,
            v.size, v.condition, v.colour, v.date.strftime('%d/%m/%Y'), f"{v.cost:.2f}", f"{v.price:.2f}", f"{v.fees:.2f}",
            f"{v.net:.2f}", f"{v.profit:.2f}", f"{v.margin:.2f}%", v.qty, v.location, v.status
        ])
    return response

@login_required
def export_xlsx(request):
    try:
        import openpyxl
    except ImportError:
        return HttpResponse("XLSX export requires openpyxl library.", status=400)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    headers = ['Main SKU','Variant SKU','Product Name','Brand','Category','Size','Condition','Colour','Date','Cost','Price','Fees','Net','Profit','Margin','Qty','Location','Status']
    ws.append(headers)
    for v in Variant.objects.select_related('product').all():
        ws.append([
            v.product.main_sku, v.variant_sku, v.product.name, v.product.brand, v.product.category,
            v.size, v.condition, v.colour, v.date.strftime('%d/%m/%Y'), float(v.cost), float(v.price), float(v.fees),
            float(v.net), float(v.profit), f"{v.margin:.2f}%", v.qty, v.location, v.status
        ])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="products.xlsx"'
    return resp

@login_required
def export_to_list_zip(request):
    to_list_qs = Variant.objects.select_related('product').prefetch_related('images').filter(status='To List', product__archived=False)
    if not to_list_qs.exists():
        messages.info(request, 'No variants with status "To List" to export.')
        return redirect('inventory:dashboard')

    mem = io.BytesIO()
    with zipfile.ZipFile(mem, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        # Root manifest for marketplaces that support CSV import
        manifest_io = io.StringIO()
        writer = csv.DictWriter(manifest_io, fieldnames=[
            'main_sku','variant_sku','name','brand','category','size','colour','condition','price','qty','location','date','images'
        ])
        writer.writeheader()

        for v in to_list_qs:
            p = v.product
            folder_name = f"{(v.variant_sku or p.main_sku) or 'SKU'}-{slugify(p.name) or 'item'}"
            base = f"{folder_name}/"

            # Description text (simple, editable after export)
            desc_parts = [
                f"{p.brand} {p.name}".strip(),
                f"Category: {p.category}",
                f"Size: {v.size}" if v.size else None,
                f"Colour: {v.colour}" if v.colour else None,
                f"Condition: {v.condition}" if v.condition else None,
                f"SKU: {v.variant_sku or p.main_sku}",
            ]
            description = "\n".join([s for s in desc_parts if s])

            meta = {
                'product_name': p.name,
                'brand': p.brand,
                'category': p.category,
                'main_sku': p.main_sku,
                'variant_sku': v.variant_sku,
                'size': v.size,
                'colour': v.colour,
                'condition': v.condition,
                'qty': v.qty,
                'price': float(v.price or 0),
                'cost': float(v.cost or 0),
                'location': v.location,
                'date': v.date.strftime('%Y-%m-%d') if v.date else '',
                'status': v.status,
            }

            # Per-item files
            zf.writestr(base + 'product.json', json.dumps(meta, indent=2))
            zf.writestr(base + 'description.txt', description)

            # Images folder
            img_count = 0
            for idx, img in enumerate(v.images.all(), start=1):
                try:
                    path = img.image.path
                except Exception:
                    continue
                try:
                    with open(path, 'rb') as fh:
                        ext = os.path.splitext(path)[1] or '.jpg'
                        zf.writestr(f"{base}images/{idx:02d}{ext}", fh.read())
                        img_count += 1
                except FileNotFoundError:
                    continue

            # Manifest row
            writer.writerow({
                'main_sku': p.main_sku,
                'variant_sku': v.variant_sku,
                'name': p.name,
                'brand': p.brand,
                'category': p.category,
                'size': v.size,
                'colour': v.colour,
                'condition': v.condition,
                'price': f"{v.price:.2f}",
                'qty': v.qty,
                'location': v.location,
                'date': v.date.strftime('%Y-%m-%d') if v.date else '',
                'images': img_count,
            })

        zf.writestr('manifest.csv', manifest_io.getvalue())

    mem.seek(0)
    resp = HttpResponse(mem.getvalue(), content_type='application/zip')
    resp['Content-Disposition'] = 'attachment; filename="to-list.zip"'
    return resp
